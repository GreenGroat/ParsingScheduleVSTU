import xlrd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, PatternFill, Color
from openpyxl.utils import get_column_letter


def convert_xls_to_xlsx(input_path):
    if input_path[-3:] == "xls":
        output_path = input_path + "x"

        # Открываем .xls файл с информацией о форматировании
        xls_book = xlrd.open_workbook(input_path, formatting_info=True)

        # Создаем новую книгу .xlsx
        xlsx_book = Workbook()
        xlsx_book.remove(xlsx_book.active)

        # Проходим по всем листам
        for sheet_index in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_index)
            xlsx_sheet = xlsx_book.create_sheet(xls_sheet.name)

            # Копируем данные и стили
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    try:
                        xls_cell = xls_sheet.cell(row, col)
                    except IndexError:
                        continue

                    xlsx_cell = xlsx_sheet.cell(row + 1, col + 1)

                    # Значение ячейки
                    value = xls_cell.value

                    # Обработка форматов данных
                    if xls_cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate.xldate_as_datetime(value, xls_book.datemode)
                        except:
                            pass

                    xlsx_cell.value = value

                    # Стили ячейки
                    if xls_cell.xf_index is not None:
                        try:
                            xf = xls_book.xf_list[xls_cell.xf_index]
                            font = xls_book.font_list[xf.font_index]

                            # Конвертация цвета шрифта
                            def get_color(color_index):
                                default_colors = {
                                    0: "000000", 1: "FFFFFF", 2: "FF0000",
                                    3: "00FF00", 4: "0000FF", 5: "FFFF00",
                                    6: "FF00FF", 7: "00FFFF"
                                }
                                if color_index in default_colors:
                                    return Color(rgb=default_colors[color_index])
                                else:
                                    rgb = xls_book.colour_map.get(color_index, (0, 0, 0))
                                    return Color(rgb=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}")

                            # Шрифт
                            xlsx_cell.font = Font(
                                name=font.name,
                                size=font.height / 20,
                                bold=font.bold,
                                italic=font.italic,
                                color=get_color(font.colour_index)
                            )

                            # Выравнивание
                            alignment = xf.alignment
                            xlsx_cell.alignment = Alignment(
                                horizontal=alignment.hor_align,
                                vertical=alignment.vert_align,
                                wrap_text=alignment.text_wrapped
                            )

                            # Границы
                            border = xf.border
                            for side in ['top', 'right', 'bottom', 'left']:
                                border_style = getattr(border, f'{side}_line_style')
                                color_index = getattr(border, f'{side}_colour_index')
                                if border_style and border_style > 0:
                                    side_obj = Side(
                                        style=convert_line_style(border_style),
                                        color=get_color(color_index)
                                    )
                                    setattr(xlsx_cell.border, side, side_obj)

                            # Заливка
                            pattern = xf.background
                            if pattern.fill_pattern > 0:
                                xlsx_cell.fill = PatternFill(
                                    patternType=convert_fill_pattern(pattern.fill_pattern),
                                    fgColor=get_color(pattern.pattern_colour_index),
                                    bgColor=get_color(pattern.background_colour_index)
                                )
                        except Exception as e:
                            # print(f"Ошибка стиля (строка {row}, колонка {col}): {str(e)}")
                            continue

            # Объединенные ячейки
            for merged in xls_sheet.merged_cells:
                xlsx_sheet.merge_cells(
                    start_row=merged[0] + 1,
                    end_row=merged[1],
                    start_column=merged[2] + 1,
                    end_column=merged[3]
                )

            # Ширина столбцов
            for col in range(xls_sheet.ncols):
                col_info = xls_sheet.colinfo_map.get(col)
                if col_info and col_info.width:
                    width = col_info.width
                    xlsx_sheet.column_dimensions[get_column_letter(col + 1)].width = width / 256 * 7 + 1

        xlsx_book.save(output_path)
        os.remove(input_path)
        return output_path


# Вспомогательные функции
def convert_line_style(style):
    styles = {
        0: None, 1: 'thin', 2: 'medium', 3: 'dashed',
        4: 'dotted', 5: 'thick', 6: 'double', 7: 'hair'
    }
    return styles.get(style)


def convert_fill_pattern(pattern):
    patterns = {
        0: 'none', 1: 'solid', 2: 'mediumGray',
        3: 'darkGray', 4: 'lightGray'
    }
    return patterns.get(pattern, 'none')
