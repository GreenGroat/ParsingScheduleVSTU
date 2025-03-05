from openpyxl import load_workbook
import inspect
from openpyxl.worksheet.merge import MergedCellRange, MergeCells
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries
from fnmatch import fnmatch
from pprint import pprint

import copy

def get_schedule_start_row_column():
    for row_id, row in enumerate(list(sheet.rows), start=1):
        for column_id, column in enumerate(list(sheet.columns)[:15], start=1):
            if sheet.cell(row=row_id, column=column_id).value and sheet.cell(row=row_id, column=column_id).value.replace(" ", "") == "1-2":
                return row_id, column_id


def is_lesson_name(text):
    if (type(text) != int) and (str(text).upper() == text) and (str(text) != 'ХХХ') and (text.count('-') == 0) and \
            len(str(text)) == len(list(letter for letter in str(text) if not(letter.isnumeric()))):
        return True
    else: return False


def week_index(row, start_values, day_len):
    return int(row > start_values[0] + 6*day_len)


def is_teacher(text):
    # Дописать получение всех преподов с бд заместо получения с json
    if type(text) == str and not all(letter == letter.upper() for letter in text) and not any(letter.isnumeric() for letter in text):
        import json
        with open("teachers.json", 'rb') as file:
            all_teachers = list(json.load(file))

        if "." in text.split()[0] and "." not in text.split()[1]:
            surname = text.split()[1]
        elif "." in text.split()[0] and "." in text.split()[1]:
            surname = text.split()[2]
        else:
            surname = text.split()[0]
        return any(surname.strip() == teacher.split()[0].strip() for teacher in all_teachers)


def is_number_audience(text):
    text = str(text).replace('-', '').replace(' ', '')
    if (
        type(text) == int or
        'зал' in (str(text).lower()) or
        len(list(letter for letter in text if letter.isnumeric())) > 2 or
        str(text).replace(",", "").replace(" ", "").isdigit()
    ) and "." not in str(text):
        return True
    else:
        return False


def sort_merged_ranges(ranges):
    # Преобразуем в кортежи с начальными строками и столбцами
    ranges_tuples = []
    for merged_range in ranges:
        start_cell = merged_range.min_row, merged_range.min_col
        ranges_tuples.append((start_cell, merged_range))

    # Сортируем по строкам и столбцам
    sorted_ranges = sorted(ranges_tuples, key=lambda x: (x[0][0], x[0][1]))

    # Возвращаем отсортированный список объектов MergedCellRange
    return [range_1[1] for range_1 in sorted_ranges]


def get_schedule_merged():
    merged_schedule = []
    start_schedule_values = get_schedule_start_row_column()
    for merge in list(sheet.merged_cells):

        if list(merge.cells)[0][0] >= start_schedule_values[0] and list(merge.cells)[0][1] > start_schedule_values[1]:
            merged_schedule.append(merge)

    return sort_merged_ranges(merged_schedule)


def check_is_the_group(value):
    all_patterns = ['[А-Я][А-Я][А-Я]-[1-9][0-9][0-9]', '[А-Я][А-Я]-[1-9][0-9][0-9]', '[А-Я][А-Я][А-Я]-[1-9][0-9][0-9]*', '[А-Я][А-Я][А-Я][А-Я]-[1-9]*', '[А-Я][А-Я]-[1-9]*', '[А-Я][А-Я][А-Я]-[1-9][0-9]*', '[А-Я][А-Я]-[1-9][0-9]*', '[А-Я][А-Я][А-Я][А-Я][1-9]', '[А-Я][А-Я][А-Я][А-Я][1-9],[А-Я][А-Я][А-Я][А-Я][1-9]*']
    for pattern in all_patterns:
        if ((type(value) is str) and fnmatch(value.replace(' ', ''), pattern)) or (str(value.strip())[0] == "Ф" and str(value.strip())[1] == "-"):
            return True


def get_group_column():
    group_column = {}
    column_group = {}  # "Развернутый" словарь group_column

    # Получаем словарик со всеми группами и номерами "их" столбцов
    for merge in list(sheet.merged_cells):
        if sheet.cell(*list(merge.cells)[0]).value and type(sheet.cell(*list(merge.cells)[0]).value) == str and check_is_the_group(sheet.cell(*list(merge.cells)[0]).value.replace(" ", "")):
            if (str(sheet.cell(*list(merge.cells)[0]).value).replace(" ", "") not in group_column.keys()) or (list(merge.cells)[0][0] < list(group_column[str(sheet.cell(*list(merge.cells)[0]).value).replace(" ", "")].cells)[0][0]):
                group_column[str(sheet.cell(*list(merge.cells)[0]).value).replace(" ", "")] = merge

    # Форматируем словарь для более удобной дальнейшей работы (Формат эквивалентен словарю с днями)
    for k, v in group_column.items():
        column_group[list(v.cells)[0][1]] = {
            "name": k,
            "range": [list(v.cells)[0][1], list(v.cells)[-1][1]]
        }

    pre_days_len = sorted(column_group.keys())
    group_len = [pre_days_len[i + 1] - pre_days_len[i] for i in range(len(pre_days_len) - 1)][0]

    return column_group, group_len


def get_schedule_days_ranges():
    start_values = get_schedule_start_row_column()
    days = {}

    months = {}
    dates = []

    # Заполнение словаря месяцев для дальнейшей работы (нужно при составлении дат)
    for month_num in range(1, start_values[0]):
        if sheet.cell(start_values[0] - 1, month_num).value:
            months[month_num] = sheet.cell(start_values[0] - 1, month_num).value

    # по примеру поиска подгрупп
    for merge in list(sheet.merged_cells):

        # Выбор всех дней недели и занесение в словарь, ключ - номер начальной строки
        if list(merge.cells)[0][1] < start_values[0] and list(merge.cells)[0][0] >= start_values[1] and sheet.cell(*list(merge.cells)[0]).value:
            if type(sheet.cell(*list(merge.cells)[0]).value) == str:
                days[list(merge.cells)[0][0]] = {
                    "name": sheet.cell(*list(merge.cells)[0]).value,
                    "range": [list(merge.cells)[0][0], list(merge.cells)[-1][0]],
                    "days": []
                }

            # Выбор всех дат
            elif type(sheet.cell(*list(merge.cells)[0]).value) == int or sheet.cell(*list(merge.cells)[0]).value.isdigit():
                dates.append((sheet.cell(*list(merge.cells)[0]).value, months[list(merge.cells)[0][1]], list(merge.cells)[0][0]))

    # Сортировка номеров строк начал дней недели для вычисления разницы между номерами строк двух смежных дней (крч, скок строчек между днями)
    pre_days_len = sorted(days.keys())
    day_len = [pre_days_len[i + 1] - pre_days_len[i] for i in range(len(pre_days_len) - 1)][0]

    # Заполнение дат в основной словарь путем вычисления начальной строки дня
    for date in dates:
        start_day_num = date[2] - date[2] % day_len + start_values[0]
        days[start_day_num if start_day_num in days.keys() else start_day_num + 1 if start_day_num + 1 in days.keys() else start_day_num - 1]["days"].append(f"{date[0]} {date[1]}")

    return days, day_len


def check_full_day(lesson_info: list):
    return any(item for item in lesson_info if is_lesson_name(item)) and (any(item for item in lesson_info if is_teacher(item))) and any(item for item in lesson_info if is_number_audience(item))


def check_num_lesson_at_day(start_values, day_len):
    lessons = set()
    for row in range(start_values[0], start_values[0] + day_len):
        if sheet.cell(row, start_values[1]).value:
            lessons.add(sheet.cell(row, start_values[1]).value.replace(" ", ""))
    return len(lessons)


def init_schedule(groups, dates, start_values, day_len):
    """
    all_schedule_type:
    { "GROUP_NAME": {
                        "NUM_WEEK": {
                                        "DAY_NUM": {
                                                        "name": str, # weekday name
                                                        "dates": list, # weekday dates on semester
                                                        "lessons":
                                                                    {"lesson_start_time":
                                                                                [
                                                                                    {
                                                                                        "lesson_name": str,
                                                                                        "lesson_audience": str,
                                                                                        "lesson_teacher": str,
                                                                                        "lesson_subgroup": int,
                                                                                        "lesson_type": str,
                                                                                        "lesson_bmt": bool,
                                                                                        "lesson_description": ""
                                                                                    }
                                                                                ...
                                                                                ]
                                                                    }
                                                    }
                                                    ...
                                        ...
                                    }
                        ...
                    }
    ...
    }
    """

    all_schedule = {}

    lessons_time_start = list(range(1, 9))
    num_lessons = check_num_lesson_at_day(start_values, day_len)

    num_lessons_times = {}
    for lesson_time in lessons_time_start[:num_lessons]:
        num_lessons_times[lesson_time] = []

    # init all_schedule
    for group_start_column in groups.keys():
        if groups[group_start_column]['name'] not in all_schedule.keys():
            all_schedule[groups[group_start_column]['name']] = {"1": {}, "2": {}}
            for k, day in enumerate(sorted(dates.keys()), start=1):
                if not all_schedule[groups[group_start_column]['name']][str(week_index(day, start_values, day_len) + 1)]:
                    all_schedule[groups[group_start_column]['name']][str(week_index(day, start_values, day_len) + 1)] = {}
                all_schedule[groups[group_start_column]['name']][str(week_index(day, start_values, day_len) + 1)][len(all_schedule[groups[group_start_column]['name']][str(week_index(day, start_values, day_len) + 1)]) + 1] = \
                    {
                     # "name": dates[day]['name'],
                     # "dates": dates[day]["days"],
                     'lessons': copy.deepcopy(num_lessons_times)}

    return all_schedule


def get_lesson_info_dict(lesson_info, lesson_type, subgroup):
    lesson_dict_info_sample = {
        "lesson_name": "",
        "lesson_audience": '',
        "lesson_teacher": '',
        "lesson_subgroup": [],
        "lesson_type": '',
        "lesson_bmt": False,
        "lesson_description": ""

    }

    for item in lesson_info:
        if is_lesson_name(item):
            lesson_dict_info_sample['lesson_name'] = str(item)
        elif is_teacher(item):
            lesson_dict_info_sample[
                'lesson_teacher'] += f"{str(item) if str(item) not in lesson_dict_info_sample['lesson_teacher'] else ''} "
        elif is_number_audience(item):
            lesson_dict_info_sample[
                'lesson_audience'] += f"{str(item) if str(item) not in lesson_dict_info_sample['lesson_audience'] else ''} "
        else:
            lesson_dict_info_sample[
                'lesson_description'] += f"{str(item) if str(item) not in lesson_dict_info_sample['lesson_description'] else ''} "

        lesson_dict_info_sample['lesson_type'] = lesson_type
        lesson_dict_info_sample['lesson_subgroup'] = subgroup


    return lesson_dict_info_sample


def get_lessons(start_values: tuple, groups: dict, group_len: int, dates: dict, day_len: int):
    lessons_at_day = check_num_lesson_at_day(start_values, day_len)

    all_schedule = init_schedule(groups, dates, start_values, day_len)
    for merge in list(sheet.merged_cells):
        lesson_info = []

        # Пробегаемся по каждой объединенной ячейке и проверяем, лежит ли она в нужном нам диапазоне
        if list(merge.cells)[0][1] >= start_values[0] and list(merge.cells)[0][0] >= start_values[1] and sheet.cell(*list(merge.cells)[0]).value and (is_lesson_name(sheet.cell(*list(merge.cells)[0]).value)) or str(sheet.cell(*list(merge.cells)[0]).value).strip() == "ОСНОВЫ ВОЕННОЙ ПОДГОТОВКИ \n4 ЧАСА":
            if str(sheet.cell(*list(merge.cells)[0]).value).strip() != "ОСНОВЫ ВОЕННОЙ ПОДГОТОВКИ \n4 ЧАСА":
                # Вычисляем последнюю возможную строку дня
                max_last_lesson_line = list(merge.cells)[0][0] - (list(merge.cells)[0][0] - start_values[0]) % day_len + day_len + week_index(list(merge.cells)[0][0], start_values, day_len)
                for row in range(list(merge.cells)[0][0], max_last_lesson_line):
                    row += week_index(row, start_values, day_len)
                    for column in range(list(merge.cells)[0][1], list(merge.cells)[-1][1] + 1):
                        if sheet.cell(row=row, column=column).value:
                            # print(column, row, sheet.cell(row=row, column=column).value)
                            lesson_info.append(sheet.cell(row=row, column=column).value)    # Записываем всю собранную информацию

                    # Если строка последняя в "паре", проверяем полноту собранной информации
                    if (row % (day_len // lessons_at_day)) == 0 and check_full_day(lesson_info):
                        # Вычисляем: строку начала дня, порядковый номер дня недели и номера пар, в которые попадает предмет (знаем строку начала пары, знаем строку конца)
                        start_day_num = row - (row - start_values[0]) % day_len + week_index(row, start_values, day_len)
                        # print(start_day_num)
                        day_num = list(k if k <= lessons_at_day else k - lessons_at_day for k, v in enumerate(sorted(dates.keys()), start=1) if v == start_day_num)[0]
                        lessons = list([k for k, l_row in enumerate(range(start_day_num, start_day_num + day_len, (day_len // lessons_at_day)), start=1) if l_row in range(list(merge.cells)[0][0], row)])
                        # print(day_num)
                        # print(lessons)
                        # print(start_day_num, day_num, days, sheet.cell(*list(merge.cells)[0]).value, list(merge.cells)[0])

                        # lecture type
                        if merge.size['columns'] > group_len:

                            # Генерируем словарь из имеющейся информации
                            lesson_dict = get_lesson_info_dict(lesson_info, "lecture", [2])

                            # Пробегаемся по всем группам и проверяем, относится ли группа к этой лекции (по номеру столбца), пока что оффнуто = (
                            for group_start_column in sorted(groups.keys()):
                                if group_start_column in range(list(merge.cells)[0][1], list(merge.cells)[-1][1] + 1):
                                    for day in lessons:
                                        if len(all_schedule[groups[group_start_column]['name']][str(week_index(row, start_values, day_len) + 1)][day_num]['lessons'][day]) - 1 not in all_schedule[groups[group_start_column]['name']][str(week_index(row, start_values, day_len) + 1)][day_num]['lessons'][day]:
                                            all_schedule[groups[group_start_column]['name']][str(week_index(row, start_values, day_len) + 1)][day_num]['lessons'][day].append(lesson_dict)

                        # practice type
                        elif merge.size['columns'] == group_len and (row - list(merge.cells)[0][0] + 1) == (day_len // lessons_at_day):

                            # Также генерируем словарь из имеющихся данных (Да-да-да, я знаю, что можно вынести в функцию, пока что так пусть будет
                            lesson_dict = get_lesson_info_dict(lesson_info, "practice", [2])

                            # Добавляем сделанный словарь в основной словарь
                            group = groups[list(merge.cells)[0][1]]['name']
                            week = str(week_index(list(merge.cells)[0][0], start_values, day_len) + 1)
                            lesson_num = lessons[0]
                            all_schedule[group][week][day_num]["lessons"][lesson_num].append(lesson_dict)

                        # all labs type
                        elif merge.size['columns'] == group_len and (row - list(merge.cells)[0][0] + 1) > (day_len // lessons_at_day):
                            lessons = list([k for k, l_row in enumerate(range(start_day_num, start_day_num + day_len, (day_len // lessons_at_day)), start=1) if l_row in range(list(merge.cells)[0][0], row)])
                            lesson_dict = get_lesson_info_dict(lesson_info, 'lab 2', [2])

                            for lesson in lessons:
                                group = groups[list(merge.cells)[0][1]]['name']
                                week = str(week_index(list(merge.cells)[0][0], start_values, day_len) + 1)
                                all_schedule[group][week][day_num]["lessons"][lesson].append(lesson_dict)

                        # subgroups labs
                        elif merge.size['columns'] == group_len // 2 and (row - list(merge.cells)[0][0] + 1) > (day_len // lessons_at_day):
                            subgroup = 0 if list(merge.cells)[0][1] in groups.keys() else 1
                            group = groups[list(merge.cells)[0][1]]['name'] if list(merge.cells)[0][1] in groups.keys() else groups[list(merge.cells)[0][1] - group_len // 2]['name']

                            lessons = list([k for k, l_row in enumerate(range(start_day_num, start_day_num + day_len, (day_len // lessons_at_day)), start=1) if l_row in range(list(merge.cells)[0][0], row)])
                            lesson_dict = get_lesson_info_dict(lesson_info, f'lab {subgroup}', [subgroup])

                            for lesson in lessons:
                                week = str(week_index(list(merge.cells)[0][0], start_values, day_len) + 1)
                                all_schedule[group][week][day_num]["lessons"][lesson].append(lesson_dict)

                        # Останавливаем цикл с поиском строк, переходим к следующему мержу
                        break



            # else:
            #     max_last_lesson_line = list(merge.cells)[0][0] - list(merge.cells)[0][0] % day_len + start_values[0] + day_len
            #     for column in range(list(merge.cells)[0][1], list(merge.cells)[-1][1] + 1):
            #         for row in range(list(merge.cells)[0][0], max_last_lesson_line):
            #             if sheet.cell(row=row, column=column).value:
            #                 lesson_info.append(sheet.cell(row=row, column=column).value)
    return all_schedule


def main():
    dates = get_schedule_days_ranges()
    # pprint(dates)
    groups = get_group_column()
    start_values = get_schedule_start_row_column()
    # pprint(groups[0])
    # print(check_num_lesson_at_day(start_values, dates[1]))
    a = get_lessons(start_values, *groups, *dates)

    # print(is_number_audience("А-404"))
    # print(is_teacher('ст. пр. Гилка В.В.'))
    # pprint(init_schedule(groups[0], dates[0], start_values, dates[1]))
    #
    import json
    with open('working.json', 'a', encoding='utf-8') as file:
        json.dump(a, file, ensure_ascii=False, indent=4)


def init_book(book_name) -> Worksheet:
    book = load_workbook(book_name)
    sheet = book.active
    return sheet


if __name__ == "__main__":
    sheet = init_book("ОН_ФЭВТ_2 курс.xlsx")
    main()
